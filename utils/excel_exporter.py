"""
Экспорт данных в Excel
"""

import os
from datetime import datetime
from typing import List, Dict, Any
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from config import EXPORT_CONFIG
from utils.logger import setup_logger


class ExcelExporter:
    """
    Класс для экспорта данных в Excel
    """

    def __init__(self, output_dir: str = None):
        """
        Инициализация экспортера

        Args:
            output_dir: Директория для сохранения файлов
        """
        self.output_dir = output_dir or EXPORT_CONFIG["output_dir"]
        self.logger = setup_logger(self.__class__.__name__)

        # Создание директории, если не существует
        os.makedirs(self.output_dir, exist_ok=True)

    def export_to_excel(
        self,
        data: List[Dict[str, Any]],
        filename: str = None,
        sheet_name: str = "Данные",
        auto_format: bool = True,
    ) -> str:
        """
        Экспорт данных в Excel файл

        Args:
            data: Список словарей с данными
            filename: Имя файла (без расширения)
            sheet_name: Название листа
            auto_format: Применять ли автоформатирование

        Returns:
            Путь к созданному файлу
        """
        if not data:
            self.logger.warning("Нет данных для экспорта")
            return None

        # Генерация имени файла
        if not filename:
            timestamp = datetime.now().strftime(EXPORT_CONFIG["date_format"])
            filename = f"export_{timestamp}"

        filepath = os.path.join(self.output_dir, f"{filename}.xlsx")

        try:
            # Создание DataFrame
            df = pd.DataFrame(data)

            # Экспорт в Excel
            self.logger.info(f"Экспорт {len(data)} записей в {filepath}")
            df.to_excel(filepath, sheet_name=sheet_name, index=False, engine="openpyxl")

            # Применение форматирования
            if auto_format:
                self._format_excel(filepath, sheet_name)

            self.logger.info(f"Данные успешно экспортированы в {filepath}")
            return filepath

        except Exception as e:
            self.logger.error(f"Ошибка при экспорте в Excel: {e}")
            return None

    def _format_excel(self, filepath: str, sheet_name: str):
        """
        Применение форматирования к Excel файлу

        Args:
            filepath: Путь к файлу
            sheet_name: Название листа
        """
        try:
            wb = load_workbook(filepath)
            ws = wb[sheet_name]

            # Стили
            header_fill = PatternFill(
                start_color="4472C4", end_color="4472C4", fill_type="solid"
            )
            header_font = Font(bold=True, color="FFFFFF", size=11)
            header_alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )

            cell_alignment = Alignment(
                horizontal="left", vertical="top", wrap_text=True
            )

            # Форматирование заголовков
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment

            # Форматирование ячеек данных
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                for cell in row:
                    cell.alignment = cell_alignment

            # Автоподбор ширины столбцов
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)

                for cell in column:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass

                adjusted_width = min(max_length + 2, 50)  # Максимум 50 символов
                ws.column_dimensions[column_letter].width = adjusted_width

            # Закрепление первой строки
            ws.freeze_panes = "A2"

            # Сохранение
            wb.save(filepath)
            self.logger.debug(f"Форматирование применено к {filepath}")

        except Exception as e:
            self.logger.error(f"Ошибка при форматировании Excel: {e}")

    def export_organizations(
        self, organizations: List[Any], filename: str = None
    ) -> str:
        """
        Экспорт списка организаций (объектов Organization)

        Args:
            organizations: Список объектов Organization
            filename: Имя файла

        Returns:
            Путь к созданному файлу
        """
        # Преобразование объектов в словари
        data = []
        for org in organizations:
            if hasattr(org, "to_dict"):
                data.append(org.to_dict())
            else:
                data.append(vars(org))

        return self.export_to_excel(
            data=data,
            filename=filename or "organizations",
            sheet_name="Аудиторские организации",
        )

    def export_multiple_sheets(
        self, sheets_data: Dict[str, List[Dict[str, Any]]], filename: str = None
    ) -> str:
        """
        Экспорт данных в Excel с несколькими листами

        Args:
            sheets_data: Словарь {название_листа: данные}
            filename: Имя файла

        Returns:
            Путь к созданному файлу
        """
        if not sheets_data:
            self.logger.warning("Нет данных для экспорта")
            return None

        # Генерация имени файла
        if not filename:
            timestamp = datetime.now().strftime(EXPORT_CONFIG["date_format"])
            filename = f"export_multi_{timestamp}"

        filepath = os.path.join(self.output_dir, f"{filename}.xlsx")

        try:
            with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
                for sheet_name, data in sheets_data.items():
                    if data:
                        df = pd.DataFrame(data)
                        df.to_excel(writer, sheet_name=sheet_name, index=False)

            self.logger.info(f"Данные успешно экспортированы в {filepath}")

            # Форматирование каждого листа
            wb = load_workbook(filepath)
            for sheet_name in sheets_data.keys():
                if sheet_name in wb.sheetnames:
                    self._format_excel(filepath, sheet_name)

            return filepath

        except Exception as e:
            self.logger.error(f"Ошибка при экспорте в Excel: {e}")
            return None
